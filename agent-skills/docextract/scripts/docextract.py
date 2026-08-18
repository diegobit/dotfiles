#!/usr/bin/env -S uv run --script
# /// script
# requires-python = ">=3.10"
# dependencies = [
#   "pymupdf>=1.24",
#   "pymupdf4llm>=0.0.17",
#   "markitdown[docx,pptx,xlsx]>=0.1",
#   "openpyxl>=3.1",
# ]
# ///
"""docextract — local document -> text / markdown / structured JSON on macOS.

Routes each input to the most accurate engine available (benchmarked, see the skill's
report): PDF text layers via PyMuPDF, image-only pages via Apple's Vision OCR,
Office formats via MarkItDown.

  docextract FILE...                 -> markdown to stdout
  docextract FILE -o out.md          -> markdown to a file
  docextract DIR -o outdir/          -> mirror the tree as .md files
  docextract FILE --json             -> structured JSON (per page, with OCR flags)
  docextract FILE --text             -> plain text, no markdown decoration
  docextract FILE --screenshot DIR   -> render pages to PNG (for visual inspection)

Exit codes: 0 ok, 1 hard failure, 3 produced no text at all.
"""
from __future__ import annotations

import argparse
import contextlib
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile

HERE = os.path.dirname(os.path.abspath(__file__))
VISION = os.path.join(HERE, 'vision_ocr')

PDF_EXT = {'.pdf'}
IMG_EXT = {'.png', '.jpg', '.jpeg', '.tif', '.tiff', '.bmp', '.gif', '.webp', '.heic'}
OFFICE_EXT = {'.docx', '.doc', '.pptx', '.ppt', '.xlsx', '.xls', '.csv', '.tsv',
              '.odt', '.odp', '.ods', '.rtf', '.epub', '.html', '.htm', '.xml', '.json'}
SUPPORTED = PDF_EXT | IMG_EXT | OFFICE_EXT

# A page with fewer than this many extractable chars is treated as needing OCR.
TEXT_LAYER_MIN_CHARS = 80
OCR_DPI = 300


def warn(msg):
    print(f'docextract: {msg}', file=sys.stderr)


@contextlib.contextmanager
def captured_stdout_fd():
    """Redirect file descriptor 1 to a temp file.

    PyMuPDF/pymupdf4llm emit progress banners from their C layer straight to fd 1, so a
    Python-level contextlib.redirect_stdout does not catch them; that noise would corrupt
    markdown written to stdout.
    """
    sys.stdout.flush()
    saved = os.dup(1)
    tmp = tempfile.TemporaryFile(mode='w+')
    try:
        os.dup2(tmp.fileno(), 1)
        yield tmp
    finally:
        sys.stdout.flush()
        os.dup2(saved, 1)
        os.close(saved)


def die(msg, code=1):
    warn(msg)
    sys.exit(code)


# --------------------------------------------------------------------------- OCR
def have_vision():
    return os.path.exists(VISION) and os.access(VISION, os.X_OK)


def build_vision():
    """Compile the Swift helper on first use; it is a build artifact, not checked in."""
    src = VISION + '.swift'
    if not os.path.exists(src):
        die(f'Vision OCR source missing at {src}')
    if not shutil.which('swiftc'):
        die('swiftc not found — install Xcode command line tools: xcode-select --install')
    warn('building the Vision OCR helper (first run only)...')
    r = subprocess.run(['swiftc', '-O', '-o', VISION, src], capture_output=True)
    if r.returncode != 0 or not have_vision():
        die(f'failed to build {VISION}:\n{r.stderr.decode()[-500:]}')


def vision_ocr(png_path, as_json=False):
    """Apple Vision OCR. Highest-accuracy OCR available on macOS, no deps, no network."""
    if not have_vision():
        build_vision()
    cmd = [VISION, png_path] + (['--json'] if as_json else [])
    r = subprocess.run(cmd, capture_output=True)
    if r.returncode != 0:
        warn(f'vision OCR failed on {png_path}: {r.stderr.decode()[-200:]}')
        return None
    return r.stdout.decode('utf-8', 'replace')


# ----------------------------------------------------------------------- helpers
def md_escape_keep(s):
    """Collapse runs of >2 blank lines; keep content untouched otherwise."""
    return re.sub(r'\n{3,}', '\n\n', s).strip()


def parse_pages_arg(spec, npages):
    """'1-3,7' -> [0,1,2,6] (0-based). None -> all."""
    if not spec:
        return list(range(npages))
    out = []
    for part in spec.split(','):
        part = part.strip()
        if not part:
            continue
        if '-' in part:
            a, b = part.split('-', 1)
            out.extend(range(int(a) - 1, int(b)))
        else:
            out.append(int(part) - 1)
    return [p for p in out if 0 <= p < npages]


def rotated_group_labels(page, min_chars=3, left_frac=0.25):
    """Recover sideways row-group labels and say which rows they cover.

    Financial tables often print a group name rotated 90 degrees in the leftmost
    column, spanning the rows it governs. Every text extractor tested either drops
    it, reverses it into single characters, or strands it in the middle of an
    unrelated row -- which actively misleads, because the rows then read as if they
    belong to the previous group. Here the label's vertical extent is intersected
    with the horizontal rows it spans, so the grouping is stated explicitly.
    """
    # Only ruled tables have row-group labels. Without this gate, rotated words inside
    # figures (an attention-visualisation page yielded 118 of them) and margin stamps
    # are mistaken for group labels, and their text then gets stripped from the body.
    if not is_table_page(page):
        return []
    left_edge = page.rect.x0 + left_frac * page.rect.width
    lines = []
    for b in page.get_text('dict').get('blocks', []):
        for ln in b.get('lines', []):
            if abs(ln.get('dir', (1, 0))[1]) > 0.5 and ln['bbox'][2] <= left_edge:
                txt = ''.join(sp['text'] for sp in ln.get('spans', [])).strip()
                if len(txt) >= min_chars:
                    lines.append((ln['bbox'], txt))
    if not lines:
        return []
    # a tall label is often typeset as several side-by-side rotated lines
    lines.sort(key=lambda t: (round(t[0][1]), t[0][0]))
    groups = []
    for bbox, txt in lines:
        merged = False
        for g in groups:
            gb = g['bbox']
            overlap = min(gb[3], bbox[3]) - max(gb[1], bbox[1])
            span = max(gb[3] - gb[1], bbox[3] - bbox[1]) or 1
            if overlap / span > 0.8 and abs(bbox[0] - gb[2]) < 30:
                g['text'] += ' ' + txt
                g['bbox'] = (min(gb[0], bbox[0]), min(gb[1], bbox[1]),
                             max(gb[2], bbox[2]), max(gb[3], bbox[3]))
                merged = True
                break
        if not merged:
            groups.append({'bbox': tuple(bbox), 'text': txt})

    # The label's text bbox is centred inside its merged cell and is much shorter
    # than it, so intersecting on the text alone under-reports the covered rows.
    # The cell's own horizontal rules are exact, so prefer them: collect the
    # borders that stop inside the label column rather than spanning the table.
    col_x0 = min(g['bbox'][0] for g in groups)
    col_x1 = max(g['bbox'][2] for g in groups)

    def is_cell_rule(x0, x1):
        # must actually cross the label column (so the neighbouring column's own row
        # separators are ignored) but stop soon after it (so full-table rules, which
        # exist on every row, are ignored too)
        return x0 <= col_x0 + 2 and x1 >= col_x1 - 2 and x1 < col_x1 + 60

    rules = set()
    for dr in page.get_drawings():
        for it in dr.get('items', []):
            if it[0] == 'l' and abs(it[1].y - it[2].y) < 1.5:
                x0, x1 = sorted((it[1].x, it[2].x))
                if is_cell_rule(x0, x1):
                    rules.add(round(it[1].y, 1))
            elif it[0] == 're':
                r = it[1]
                if is_cell_rule(r.x0, r.x1):
                    rules.add(round(r.y0, 1))
                    rules.add(round(r.y1, 1))
    rules = sorted(rules)

    def cell_band(bbox):
        """Widen a label bbox to the rules immediately above and below it."""
        above = [y for y in rules if y <= bbox[1] + 1]
        below = [y for y in rules if y >= bbox[3] - 1]
        return (above[-1] if above else bbox[1], below[0] if below else bbox[3])

    words = [w for w in page.get_text('words') if abs(w[1] - w[3]) < 40]
    out = []
    for g in groups:
        x1 = g['bbox'][2]
        y0, y1 = cell_band(g['bbox'])
        keys = []
        for w in sorted(words, key=lambda w: (w[1], w[0])):
            mid = (w[1] + w[3]) / 2
            if y0 - 1 <= mid <= y1 + 1 and w[0] >= x1 - 2:
                band = round(mid)
                if not keys or abs(band - keys[-1][0]) > 3:
                    keys.append((band, w[4]))
        if keys and rules:
            out.append({'label': ' '.join(g['text'].split()),
                        'covers': [k for _, k in keys]})
    return out


def image_regions(page, min_pixels=120_000, min_area_frac=0.15,
                  aspect_range=(0.12, 8.0), max_regions=8):
    """Raster images big enough to plausibly carry readable text, as page rectangles.

    Gating on page-area fraction alone is wrong: a matrix pasted into a table cell can
    cover only ~3% of the page and still hold a full table, while a header logo covers a
    similar fraction and holds nothing. Pixel count separates them (a logo here is ~28k
    pixels, the smallest real content image ~214k), so either test can admit a region.
    """
    page_area = abs(page.rect.width * page.rect.height) or 1.0
    rects, seen = [], set()
    for img in page.get_images(full=True):
        xref, w, h = img[0], img[2], img[3]
        big_enough = (w * h) >= min_pixels
        ratio = (w / h) if h else 0
        if not (aspect_range[0] <= ratio <= aspect_range[1]):
            continue
        try:
            for r in page.get_image_rects(xref):
                if not big_enough and abs(r.width * r.height) / page_area < min_area_frac:
                    continue
                key = (round(r.x0), round(r.y0), round(r.x1), round(r.y1))
                if key in seen:
                    continue
                seen.add(key)
                rects.append(r)
        except Exception:  # noqa: BLE001 - a broken xref must not kill extraction
            continue
    if len(rects) > max_regions:
        warn(f'page has {len(rects)} large images; OCR-ing the {max_regions} biggest')
        rects.sort(key=lambda r: -abs(r.width * r.height))
        rects = rects[:max_regions]
    return rects


def new_lines_only(ocr_text, known_tokens, min_new=0.4):
    """Keep OCR lines that are not already covered by the page's text layer.

    A full-page image sits under the text layer, so OCR-ing it returns the page's own
    words again plus whatever is baked into the picture. Only the latter is worth adding.
    """
    out = []
    for line in ocr_text.split('\n'):
        toks = [t for t in re.findall(r"[0-9A-Za-zÀ-ÿ']+", line.lower()) if len(t) > 2]
        if not toks:
            continue
        fresh = sum(1 for t in toks if t not in known_tokens)
        if fresh / len(toks) >= min_new:
            out.append(line.strip())
    return out


def has_stacked_header(page, band=14, min_hits=6):
    """Detect a multi-level column header: a spanning row above finer sub-columns.

    Slice the page horizontally and count how many vertical rules cross each slice; a
    slice with noticeably FEWER column boundaries sitting directly above one with more
    is a header cell spanning sub-columns. Measured on this corpus: 12 hits on the
    stacked-header page versus 3, 1 and 1 on a simple 3-column table, a financial
    statement and a prose page. Heuristic, hence the margin.
    """
    verts = []
    for d in page.get_drawings():
        for it in d.get('items', []):
            if it[0] == 'l' and abs(it[1].x - it[2].x) < 1.5 and abs(it[1].y - it[2].y) > 3:
                verts.append((round(it[1].x, 1), min(it[1].y, it[2].y), max(it[1].y, it[2].y)))
            elif it[0] == 're' and it[1].height > 3:
                r = it[1]
                verts += [(round(r.x0, 1), r.y0, r.y1), (round(r.x1, 1), r.y0, r.y1)]
    if len(verts) < 6:
        return False
    counts, y = [], page.rect.y0
    while y < page.rect.y1:
        counts.append(len({x for x, a, b in verts if a <= y + band / 2 <= b}))
        y += band
    hits = sum(1 for i in range(len(counts) - 1)
               if 2 <= counts[i] < counts[i + 1] and counts[i + 1] - counts[i] >= 2)
    return hits >= min_hits


def probe(path):
    """Report what a PDF contains so the caller can pick a mode from data, not a guess."""
    import pymupdf
    doc = pymupdf.open(path)
    info = {'file': path, 'pages': doc.page_count, 'scanned_pages': [], 'table_pages': [],
            'stacked_header_pages': [], 'row_group_pages': [], 'figure_pages': [],
            'image_text_pages': []}
    for i in range(doc.page_count):
        pg = doc[i]
        n = len((pg.get_text(sort=True) or '').strip())
        if n < TEXT_LAYER_MIN_CHARS:
            info['scanned_pages'].append(i + 1)
            continue
        if is_table_page(pg):
            info['table_pages'].append(i + 1)
            if has_stacked_header(pg):
                info['stacked_header_pages'].append(i + 1)
        if rotated_group_labels(pg):
            info['row_group_pages'].append(i + 1)
        if count_figures(pg):
            info['figure_pages'].append(i + 1)
        if image_regions(pg):
            info['image_text_pages'].append(i + 1)
    doc.close()

    # The recommendation is per page, not per document: a single file often mixes
    # row-per-record tables (which want --text) with stacked-header tables (which want
    # markdown), so a single document-level verdict would be wrong for half of it.
    why = []
    if info['scanned_pages']:
        why.append(f'{len(info["scanned_pages"])} page(s) have no text layer; those are OCR\'d '
                   f'with Apple Vision either way')
    if info['table_pages']:
        why.append(f'{len(info["table_pages"])} ruled table page(s): --text keeps every row '
                   f'bound to its own value (cell binding 1.00)')
    if info['stacked_header_pages']:
        why.append(f'pages {info["stacked_header_pages"][:8]} look like they have stacked '
                   f'multi-level headers, where markdown scored better (18/20 vs 9/20). This is '
                   f'a heuristic and does over-flag (a wrapped single-level header can trip it) '
                   f'-- confirm with --screenshot before re-extracting those pages')
    if info['row_group_pages']:
        why.append(f'pages {info["row_group_pages"][:8]} have sideways row-group labels; read '
                   f'the row-group comments before attributing a row to a group')
    if info['figure_pages']:
        why.append(f'pages {info["figure_pages"][:8]} hold figures; no text mode can answer a '
                   f'question about those, use --screenshot and look')
    if info['image_text_pages']:
        why.append(f'pages {info["image_text_pages"][:8]} carry a large image that may have text '
                   f'baked in; add --image-ocr to recover it')
    info['recommend'] = '--text'
    info['then_recheck_pages_with_markdown'] = info['stacked_header_pages']
    info['because'] = why
    return info


def strip_rotated_fragments(text, groups):
    """Remove sideways-label glyphs that landed inside unrelated table rows.

    A rotated label is read as ordinary words, so it gets interleaved into whichever
    rows it passes: "22% EUR 116.181,94  DISPOSIZIONE  B3  Corrispettivo OIS". The
    grouping is already stated in the row-group comment, so drop the stray fragments
    rather than leave them to be misread as row content. Matching is case-sensitive on
    whole words, so ordinary prose ("Importo Somme a disposizione") is untouched.
    """
    frags = {w for g in groups for w in g['label'].split() if len(w) > 2}
    if not frags:
        return text
    out = []
    for line in text.split('\n'):
        kept = [w for w in line.split(' ') if w.strip() not in frags]
        new = ' '.join(kept)
        out.append(new if new.strip() else line if not line.strip() else new)
    return re.sub(r'[ \t]{2,}', lambda m: m.group(0), '\n'.join(out))


def is_table_page(page, min_rules=8):
    """True when the page carries enough horizontal rules to be a ruled table."""
    n = 0
    for d in page.get_drawings():
        for it in d.get('items', []):
            if it[0] == 'l' and abs(it[1].y - it[2].y) < 1.5:
                n += 1
            elif it[0] == 're':
                n += 2
    return n >= min_rules


def last_heading(page, min_len=4, max_len=80):
    """The last section heading on a page, by font size relative to the body text.

    Used to carry section context across a page break: a table continuing onto the next
    page repeats neither its header nor its section banner, so the continuation rows are
    otherwise unattributable when that page is read on its own.
    """
    spans = []
    for b in page.get_text('dict').get('blocks', []):
        for ln in b.get('lines', []):
            for sp in ln.get('spans', []):
                txt = sp['text'].strip()
                if txt:
                    spans.append((ln['bbox'][1], sp.get('size', 0), txt))
    if not spans:
        return None
    sizes = sorted(sp[1] for sp in spans)
    body = sizes[len(sizes) // 2]
    heads = [sp for sp in spans
             if sp[1] > body + 0.4 and min_len <= len(sp[2]) <= max_len]
    return max(heads, key=lambda sp: sp[0])[2] if heads else None


def count_figures(page, min_area_frac=0.03):
    """Rough count of things on a page that only vision can interpret.

    Two signals, both size-gated so page furniture does not trip them:
    raster images covering at least `min_area_frac` of the page (a header logo is
    ~1%, so it is ignored), and vector paths with more than a few line segments --
    charts are drawn as polylines, whereas table rules are single lines or rectangles.
    Measured on this corpus: 1 on the page holding a line chart, 0 on four dense
    table pages.
    """
    page_area = abs(page.rect.width * page.rect.height) or 1.0
    n = 0
    for img in page.get_images(full=True):
        try:
            for r in page.get_image_rects(img[0]):
                if abs(r.width * r.height) / page_area >= min_area_frac:
                    n += 1
                    break
        except Exception:  # noqa: BLE001 - a broken xref must not kill extraction
            continue
    for d in page.get_drawings():
        items = d.get('items', [])
        if any(it[0] == 'c' for it in items) or sum(1 for it in items if it[0] == 'l') > 4:
            r = d.get('rect')
            if r is None or abs(r.width * r.height) / page_area >= min_area_frac:
                n += 1
    return n


# --------------------------------------------------------------------------- PDF
def extract_pdf(path, pages_spec=None, force_ocr=False, want='md', image_ocr=False):
    """Returns (list_of_page_dicts, meta). Each page: {page, text, engine, chars}."""
    try:
        import pymupdf
    except ImportError:
        die('PyMuPDF is required: pip install pymupdf pymupdf4llm')

    doc = pymupdf.open(path)
    idxs = parse_pages_arg(pages_spec, doc.page_count)
    if not idxs:
        die(f'no pages selected from {path} (document has {doc.page_count})')

    # classify each page
    text_pages, ocr_pages, chars, img_pages = [], [], {}, []
    for i in idxs:
        n = 0 if force_ocr else len((doc[i].get_text(sort=True) or '').strip())
        chars[i] = n
        (ocr_pages if n < TEXT_LAYER_MIN_CHARS else text_pages).append(i)

    md_by_page = {}
    if text_pages and want == 'md':
        try:
            import pymupdf4llm
            # pymupdf4llm chats to stdout ("Using Tesseract for OCR processing."), which
            # would corrupt our markdown; capture it and re-emit on stderr.
            with captured_stdout_fd() as buf:
                chunks = pymupdf4llm.to_markdown(doc, pages=text_pages, page_chunks=True)
                buf.seek(0)
                noise = buf.read().strip()
            if noise:
                warn('pymupdf4llm: ' + ' '.join(noise.split())[:200])
            # 1.28's page_chunks metadata carries no page number, but chunks come back
            # in the order we requested, so zip them back onto text_pages.
            for pno, chunk in zip(text_pages, chunks):
                md_by_page[pno] = chunk.get('text', '')
        except ImportError:
            warn('pymupdf4llm not installed; falling back to plain text for text pages')
        except Exception as e:  # noqa: BLE001 - never let markdown extras break extraction
            warn(f'pymupdf4llm failed ({e}); falling back to plain text')

    def table_header_of(md):
        """Header + separator of the last *informative* markdown table in `md`.

        Skips banner rows (a single spanning cell), which carry no column names.
        """
        lines = md.split('\n')
        fallback = None
        for i in range(len(lines) - 1, 0, -1):
            if re.match(r'^\s*\|[\s:|-]+\|\s*$', lines[i]) and lines[i - 1].strip().startswith('|'):
                cells = [c for c in lines[i - 1].strip().strip('|').split('|') if c.strip()]
                if len(cells) > 1:
                    return lines[i - 1], lines[i]
                fallback = fallback or (lines[i - 1], lines[i])
        return fallback

    def continues_table(prev, cur, want):
        """Does `cur` look like the continuation of a table that ended `prev`?"""
        if want == 'md':
            return cur.lstrip().startswith('|') and prev['text'].rstrip().endswith('|')
        prev_rows = [l for l in prev['text'].split('\n')[-4:] if len(l.split()) > 2]
        cur_rows = [l for l in cur.split('\n')[:4] if len(l.split()) > 2]
        return bool(prev_rows and cur_rows and prev.get('is_table') and
                    re.search(r'\d', cur_rows[0]))

    out = []
    tmpdir = None
    try:
        for i in idxs:
            if i in ocr_pages:
                if tmpdir is None:
                    tmpdir = tempfile.mkdtemp(prefix='docextract_')
                png = os.path.join(tmpdir, f'p{i+1}.png')
                zoom = OCR_DPI / 72.0
                doc[i].get_pixmap(matrix=pymupdf.Matrix(zoom, zoom)).save(png)
                t = vision_ocr(png) or ''
                out.append({'page': i + 1, 'text': t.strip(), 'engine': 'vision-ocr',
                            'text_layer_chars': chars[i], 'figures': count_figures(doc[i])})
            else:
                t = md_by_page.get(i)
                if t is None:
                    # sort=True lays words out in reading order and keeps columns
                    # aligned; the unsorted default emits raw block order, which
                    # separates a table row's label from its value (measured: cell
                    # binding 0.14 unsorted vs 1.00 sorted on real fee tables).
                    t = doc[i].get_text(sort=True) or ''
                # A table continuing across a page break repeats neither its header nor
                # its section banner, so state both. In markdown this also matters
                # structurally: pymupdf4llm restarts the table per page, which promotes
                # the continuation's first DATA row to a header.
                if t and out and continues_table(out[-1], t, want):
                    prev = out[-1]
                    sect = prev.get('last_heading')
                    marker = (f'<!-- table continues from page {prev["page"]}'
                              + (f'; section: "{sect}"' if sect else '') + ' -->')
                    if want == 'md':
                        hdr = table_header_of(prev['text'])
                        body = t.lstrip().split('\n')
                        if hdr and len(body) > 1 and re.match(r'^\s*\|[\s:|-]+\|\s*$', body[1]):
                            # body[0] is a data row and body[1] the separator that
                            # wrongly promoted it; drop it and supply the real header.
                            t = (marker + '\n' + hdr[0] + '\n' + hdr[1] + '\n'
                                 + '\n'.join([body[0]] + body[2:]))
                        else:
                            t = marker + '\n' + t
                    else:
                        t = marker + '\n' + t
                # A page can carry a text layer AND text baked into a picture; the text
                # layer alone silently loses the latter. OCR just the image regions and
                # append whatever the text layer does not already cover.
                regions = image_regions(doc[i])
                if regions:
                    img_pages.append(i + 1)
                if regions and t and image_ocr:
                    known = set(re.findall(r"[0-9a-zà-ÿ']+", t.lower()))
                    extra = []
                    for n, r in enumerate(regions):
                        if tmpdir is None:
                            tmpdir = tempfile.mkdtemp(prefix='docextract_')
                        png = os.path.join(tmpdir, f'img{i+1}_{n}.png')
                        z = OCR_DPI / 72.0
                        doc[i].get_pixmap(matrix=pymupdf.Matrix(z, z), clip=r).save(png)
                        got = vision_ocr(png) or ''
                        extra += new_lines_only(got, known)
                    if extra:
                        t += ('\n\n<!-- text found inside an image on this page (OCR) -->\n'
                              + '\n'.join(extra))
                groups = rotated_group_labels(doc[i])
                if groups:
                    t = strip_rotated_fragments(t, groups)
                    note = '\n'.join(
                        f'<!-- row-group "{g["label"]}" covers rows: '
                        f'{", ".join(g["covers"])} -->' for g in groups)
                    t = note + '\n\n' + t
                out.append({'page': i + 1, 'text': t.strip(),
                            'engine': 'pymupdf4llm' if i in md_by_page else 'pymupdf',
                            'text_layer_chars': chars[i], 'figures': count_figures(doc[i]),
                            'row_groups': groups, 'last_heading': last_heading(doc[i]),
                            'is_table': is_table_page(doc[i])})
    finally:
        if tmpdir:
            shutil.rmtree(tmpdir, ignore_errors=True)

    meta = {'pages_total': doc.page_count, 'pages_done': len(out),
            'ocr_pages': [p + 1 for p in ocr_pages],
            'figure_pages': [p['page'] for p in out if p.get('figures')],
            'image_text_pages': sorted(set(img_pages))}
    doc.close()
    return out, meta


# ------------------------------------------------------------------------ images
def extract_image(path):
    t = vision_ocr(path) or ''
    return ([{'page': 1, 'text': t.strip(), 'engine': 'vision-ocr',
              'text_layer_chars': 0, 'figures': 1}],
            {'pages_total': 1, 'pages_done': 1, 'ocr_pages': [1], 'figure_pages': [1]})


# ------------------------------------------------------------------------ office
XLSX_EXT = {'.xlsx', '.xlsm'}


def _md_cell(v):
    if v is None:
        return ''
    return str(v).replace('|', '\\|').replace('\n', ' ').strip()


def extract_xlsx(path):
    """Read .xlsx natively so merged cells can be expanded.

    Every other tool mishandles merges: pandas-backed readers leave the covered cells
    as NaN (ambiguous with a genuinely empty cell), and spatial extractors place the
    label at the merge's visual centre, which silently binds it to the WRONG row.
    Here each covered cell repeats its anchor's value, so every row stands alone and
    stacked headers stay aligned to their columns.
    """
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    parts = []
    for ws in wb.worksheets:
        if ws.max_row is None or ws.max_row < 1:
            continue
        grid = [[c.value for c in row] for row in ws.iter_rows()]
        for rng in ws.merged_cells.ranges:
            anchor = grid[rng.min_row - 1][rng.min_col - 1]
            for r in range(rng.min_row - 1, rng.max_row):
                for c in range(rng.min_col - 1, rng.max_col):
                    grid[r][c] = anchor
        rows = [[_md_cell(v) for v in r] for r in grid]
        while rows and not any(c for c in rows[-1]):
            rows.pop()
        if not rows:
            continue
        width = max(len(r) for r in rows)
        rows = [r + [''] * (width - len(r)) for r in rows]
        parts.append(f'## {ws.title}\n')
        parts.append('| ' + ' | '.join(rows[0]) + ' |')
        parts.append('|' + '---|' * width)
        for r in rows[1:]:
            parts.append('| ' + ' | '.join(r) + ' |')
        parts.append('')
    return '\n'.join(parts)


def extract_office(path):
    """xlsx natively (for merge fidelity); everything else via MarkItDown."""
    ext = os.path.splitext(path)[1].lower()
    if ext in XLSX_EXT:
        try:
            txt = extract_xlsx(path)
            return ([{'page': 1, 'text': txt.strip(), 'engine': 'openpyxl',
                      'text_layer_chars': len(txt)}],
                    {'pages_total': 1, 'pages_done': 1, 'ocr_pages': []})
        except ImportError:
            warn('openpyxl not installed; falling back to MarkItDown for this workbook')
        except Exception as e:  # noqa: BLE001 - fall back rather than lose the file
            warn(f'openpyxl failed ({e}); falling back to MarkItDown')
    try:
        from markitdown import MarkItDown
    except ImportError:
        die("MarkItDown is required for Office formats: pip install 'markitdown[all]'")
    txt = MarkItDown().convert(path).text_content
    return ([{'page': 1, 'text': (txt or '').strip(), 'engine': 'markitdown',
              'text_layer_chars': len(txt or '')}],
            {'pages_total': 1, 'pages_done': 1, 'ocr_pages': []})


# --------------------------------------------------------------------- top level
def extract(path, pages=None, force_ocr=False, want='md', image_ocr=False):
    ext = os.path.splitext(path)[1].lower()
    if ext in PDF_EXT:
        return extract_pdf(path, pages, force_ocr, want, image_ocr)
    if ext in IMG_EXT:
        return extract_image(path)
    if ext in OFFICE_EXT:
        return extract_office(path)
    die(f'unsupported extension {ext!r} for {path}')


def render_text(pages, want, with_page_marks):
    parts = []
    for p in pages:
        if not p['text']:
            continue
        if with_page_marks:
            fig = p.get('figures') or 0
            note = (f' — {fig} figure(s), inspect with --screenshot' if fig else '')
            parts.append(f'<!-- page {p["page"]} ({p["engine"]}){note} -->')
        parts.append(p['text'])
    body = '\n\n'.join(parts)
    if want == 'text':
        body = re.sub(r'^\s{0,3}#{1,6}\s*', '', body, flags=re.M)
        body = re.sub(r'[*_`]{1,3}', '', body)
    return md_escape_keep(body)


def screenshot(path, outdir, pages_spec=None, dpi=150):
    import pymupdf
    os.makedirs(outdir, exist_ok=True)
    doc = pymupdf.open(path)
    made = []
    for i in parse_pages_arg(pages_spec, doc.page_count):
        z = dpi / 72.0
        dst = os.path.join(outdir, f'page_{i+1}.png')
        doc[i].get_pixmap(matrix=pymupdf.Matrix(z, z)).save(dst)
        made.append(dst)
    doc.close()
    return made


def iter_inputs(targets, recursive=True):
    for t in targets:
        if os.path.isdir(t):
            for root, _dirs, files in os.walk(t):
                for f in sorted(files):
                    if os.path.splitext(f)[1].lower() in SUPPORTED:
                        yield os.path.join(root, f)
                if not recursive:
                    break
        elif os.path.isfile(t):
            yield t
        else:
            warn(f'not found: {t}')


def _one_file(job):
    """Extract one file and write its markdown. Runs in a worker process."""
    src, indir, outdir, want, pages_spec, force_ocr, image_ocr, marks = job
    try:
        page_list, meta = extract(src, pages_spec, force_ocr, want, image_ocr=image_ocr)
    except SystemExit as e:
        return src, None, f'exit {e.code}'
    except Exception as e:  # noqa: BLE001 - one bad file must not kill the batch
        return src, None, str(e)
    body = render_text(page_list, want, marks)
    rel = os.path.relpath(src, indir)
    dst = os.path.join(outdir, os.path.splitext(rel)[0] + '.md')
    os.makedirs(os.path.dirname(dst), exist_ok=True)
    with open(dst, 'w', encoding='utf-8') as f:
        f.write(f'# {os.path.basename(src)}\n\n{body}\n')
    return src, {'dst': dst, 'chars': sum(len(p['text']) for p in page_list),
                 'ocr': len(meta.get('ocr_pages', []))}, None


def main():
    ap = argparse.ArgumentParser(
        prog='docextract',
        description='Extract text/markdown/JSON from documents locally (macOS).')
    ap.add_argument('inputs', nargs='+', help='files and/or directories')
    ap.add_argument('-o', '--output', help='output file, or output DIR when input is a dir')
    ap.add_argument('--json', action='store_true', help='structured JSON instead of markdown')
    ap.add_argument('--text', action='store_true', help='plain text instead of markdown')
    ap.add_argument('--pages', help='page selection, e.g. "1-5,10" (PDF only)')
    ap.add_argument('--force-ocr', action='store_true',
                    help='OCR every page even if a text layer exists')
    ap.add_argument('--image-ocr', action='store_true',
                    help='also OCR large images sitting on pages that DO have a text layer, '
                         'to recover text baked into a picture (off by default: it adds '
                         'lower-confidence text beside the exact text layer)')
    ap.add_argument('--probe', action='store_true',
                    help='report what the document contains and which mode suits it, '
                         'then exit (cheap; reads no page twice)')
    ap.add_argument('--screenshot', metavar='DIR',
                    help='render pages to PNG in DIR instead of extracting')
    ap.add_argument('--dpi', type=int, default=150, help='screenshot DPI (default 150)')
    ap.add_argument('--no-page-marks', action='store_true',
                    help='omit the "<!-- page N -->" comments')
    ap.add_argument('--no-recursive', action='store_true', help='do not walk subdirectories')
    ap.add_argument('-j', '--jobs', type=int, default=0,
                    help='parallel workers for directory mode '
                         '(default: CPU count minus 2; 1 disables)')
    args = ap.parse_args()

    want = 'json' if args.json else ('text' if args.text else 'md')

    if args.probe:
        for src in iter_inputs(args.inputs, not args.no_recursive):
            if os.path.splitext(src)[1].lower() not in PDF_EXT:
                continue
            print(json.dumps(probe(src), indent=1, ensure_ascii=False))
        return

    if args.screenshot:
        for src in iter_inputs(args.inputs, not args.no_recursive):
            if os.path.splitext(src)[1].lower() not in PDF_EXT:
                continue
            made = screenshot(src, args.screenshot, args.pages, args.dpi)
            print(f'{src}: {len(made)} page(s) -> {args.screenshot}')
        return

    files = list(iter_inputs(args.inputs, not args.no_recursive))
    if not files:
        die('no supported input files found')

    dir_mode = len(args.inputs) == 1 and os.path.isdir(args.inputs[0]) and args.output
    produced_any = False
    results = []

    # Directory work is CPU-bound and per-file independent, so fan it out. Only for
    # markdown/text output: --json accumulates one document per record in order.
    if dir_mode and want != 'json' and len(files) > 1:
        jobs = args.jobs or max(1, (os.cpu_count() or 2) - 2)
        if jobs > 1:
            import concurrent.futures as cf
            payload = [(f, args.inputs[0], args.output, want, args.pages, args.force_ocr,
                        args.image_ocr, not args.no_page_marks) for f in files]
            done = 0
            with cf.ProcessPoolExecutor(max_workers=jobs) as pool:
                for src, ok, err in pool.map(_one_file, payload):
                    done += 1
                    if err:
                        warn(f'FAILED {src}: {err}')
                    elif ok['chars']:
                        produced_any = True
                        print(f'{src} -> {ok["dst"]}  ({ok["chars"]} chars, '
                              f'{ok["ocr"]} OCR page(s))')
                    else:
                        warn(f'no text extracted from {src}')
            print(f'{done} file(s) with {jobs} workers', file=sys.stderr)
            if not produced_any:
                sys.exit(3)
            return

    for src in files:
        try:
            pages, meta = extract(src, args.pages, args.force_ocr, want,
                                  image_ocr=args.image_ocr)
        except SystemExit:
            raise
        except Exception as e:  # noqa: BLE001 - one bad file must not kill a batch
            warn(f'FAILED {src}: {e}')
            continue
        total_chars = sum(len(p['text']) for p in pages)
        if total_chars:
            produced_any = True
        else:
            warn(f'no text extracted from {src}')

        if want == 'json':
            results.append({'file': src, **meta, 'chars': total_chars, 'pages': pages})
            continue

        body = render_text(pages, want, not args.no_page_marks)
        if dir_mode:
            rel = os.path.relpath(src, args.inputs[0])
            dst = os.path.join(args.output, os.path.splitext(rel)[0] + '.md')
            os.makedirs(os.path.dirname(dst), exist_ok=True)
            with open(dst, 'w', encoding='utf-8') as f:
                f.write(f'# {os.path.basename(src)}\n\n{body}\n')
            print(f'{src} -> {dst}  ({total_chars} chars, '
                  f'{len(meta["ocr_pages"])} OCR page(s))')
        elif args.output and len(files) == 1:
            with open(args.output, 'w', encoding='utf-8') as f:
                f.write(body + '\n')
            print(f'{src} -> {args.output}  ({total_chars} chars, '
                  f'{len(meta["ocr_pages"])} OCR page(s))', file=sys.stderr)
        else:
            if len(files) > 1:
                print(f'\n<!-- ===== {src} ===== -->\n')
            print(body)

    if want == 'json':
        out = json.dumps(results if len(results) != 1 else results[0],
                         indent=1, ensure_ascii=False)
        if args.output:
            with open(args.output, 'w', encoding='utf-8') as f:
                f.write(out + '\n')
            print(f'-> {args.output}', file=sys.stderr)
        else:
            print(out)

    if not produced_any:
        sys.exit(3)


if __name__ == '__main__':
    main()
